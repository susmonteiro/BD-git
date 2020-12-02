import openpyxl, random
outSql = open("populate.sql", "w")

#query 3
allAroucaFarm = []

#init datas
allDatas = []
# this month
for day in range(1, 31):
    allDatas.append("2020-11-" + str(day))

for mes in range(1, 7):
    for day in range(1, 32):
        allDatas.append("2019-" + str(mes) + "-" + str(day))




# Regioes
allNumR = [1, 2, 3, 4, 5]
outSql.write("-- Regioes\n")
outSql.write("insert into regiao values (1, 'Norte', 3573000);\n")
outSql.write("insert into regiao values (2, 'Centro', 2217000);\n")
outSql.write("insert into regiao values (3, 'Lisboa', 4457358);\n")
outSql.write("insert into regiao values (4, 'Alentejo', 234284);\n")
outSql.write("insert into regiao values (5, 'Algarve', 438864);\n")

# Concelhos
outSql.write("\n-- Concelhos\n")
wb = openpyxl.load_workbook('concelhos.xlsx', data_only=True)
allNumPlaces = []
allNumCedula = []
allNumDoente = []
allFarmacias = []
for sName in wb.sheetnames:
    sheet = wb.get_sheet_by_name(sName)
    nRegiao = wb.sheetnames.index(sName)
    # [NomeConcelho, NumConcelho]
    if sName == "Folha1":
        continue
    for l in range(1, sheet.max_row+1):
        cell = "D" + str(l) 
        nomeC = sheet[cell].value.title()
        cell = "E" + str(l) 
        numC = sheet[cell].value
        tpl = (nRegiao, numC)
        allNumPlaces.append(tpl)

        # insert into concelho values (numC, numReg, nomeC, numHab)
        outSql.write("insert into concelho values (" + \
            str(numC) + ", " + str(nRegiao) + ", '" + str(nomeC).strip() + "', " + str(random.randint(10000, 1000000)) + ");\n")
wb.close()

# Instituicoes
outSql.write("\n-- Instituicoes\n")
instFile = open("inst.txt", "r")
allInstTypes = ['farmacia', 'laboratorio', 'clinica', 'hospital']
counter = 1
AroucaPlace = (str(2), str(104))
while instFile:
    line  = instFile.readline()
    counter += 1
    if line == "":
        break
    tipo = random.choice(allInstTypes)
    nome = line[:-1]
    tpl = random.choice(allNumPlaces)
    if tipo == 'farmacia':
        allFarmacias.append(nome)
        if counter % 5 == 0:
            allAroucaFarm.append(nome)
            tpl = AroucaPlace
    # insert into instituicao values (nome, tipo, num_regiao, num_concelho)
    outSql.write("insert into instituicao values ('" + nome + "', '" + \
        tipo + "', " +  str(tpl[0]) + ", " + str(tpl[1]) + ");\n")

instFile.close()


# Medico
outSql.write("\n-- Medicos\n")
namesFile = open("names.txt", "r")
nameLines = namesFile.readlines()
especialidadeFile = open("especialidades.txt", "r")
especialidadeLines = especialidadeFile.readlines()
for med in range(1, 150):
    # num_cedula = str(random.randint(1000, 9999))
    num_cedula = str(med)
    allNumCedula.append(num_cedula)
    outSql.write("insert into medico values (" + num_cedula + ", '" + random.choice(nameLines)[:-1] + "', '" + random.choice(especialidadeLines)[:-1] + "');\n")
    
namesFile.close()
especialidadeFile.close()


# Consulta - num_cedula, num_doente, data, nome_instituicao
outSql.write("\n-- Consulta\n")
instFile = open("inst.txt", "r")
instLines = instFile.readlines()
allConsulta = []

for i in range(1, 500):
    num_doente = str(i)
    allNumDoente.append(num_doente)

    tpl = (str(random.choice(allNumCedula)), num_doente, random.choice(allDatas))
    allConsulta.append(tpl)

    outSql.write("insert into consulta values (" + tpl[0] + ", " + tpl[1] + ", '" + tpl[2] + "', '" + random.choice(instLines)[:-1] + "');\n")

for i in range(550):
    tpl = (str(random.choice(allNumCedula)), random.choice(allNumDoente), random.choice(allDatas))
    allConsulta.append(tpl)

    outSql.write("insert into consulta values (" + tpl[0] + ", " + tpl[1] + ", '" + tpl[2] + "', '" + random.choice(instLines)[:-1] + "');\n")
instFile.close()


# query 3
allArouca = []
for i in range(3):
    num_cedula = random.choice(allNumCedula)
    for f in allAroucaFarm:
        num_doente = random.choice(allNumDoente)
        data = '2020-11-18'
        lst = [num_cedula, num_doente, data, f]
        allArouca.append(lst)
        outSql.write("insert into consulta values (" + lst[0] + ", " + lst[1] + ", '" + lst[2] + "', '" + f + "');\n")



# Prescricao
outSql.write("\n-- Prescricao\n")
medFile = open('med.txt', "r")
medLines = medFile.readlines()
allPrescricao = []
for i in range(500):
    quant = random.randint(1,9)
    subs = random.choice(medLines)[:-1]
    if i % 9 == 0:
        subs = 'Aspirina'
    tpl = random.choice(allConsulta)
    tpl_new = list(tpl) + [subs, quant]
    allPrescricao.append(tpl_new)
    outSql.write("insert into prescricao values (" + tpl[0] + ", " + tpl[1] + ", '" + tpl[2] + "', '" + subs + "', " + str(quant) + ");\n")
medFile.close()


for tpl in allArouca:
    subs = "Aspirina"
    quant = random.randint(1,9)
    outSql.write("insert into prescricao values (" + tpl[0] + ", " + tpl[1] + ", '" + tpl[2] + "', '" + subs + "', " + str(quant) + ");\n")


# Analise 
outSql.write("\n-- Analise\n")
instFile = open("inst.txt", "r")
analiseFile = open("analises.txt", "r")
analiseLines = analiseFile.readlines()
instLines = instFile.readlines()
for i in range(1, 800):
    # num_analise = str(random.randint(1000, 9999))
    num_analise = str(i) 
    tpl = random.choice(allConsulta)

    if i % 8 == 0:
        nome = "Glicémia"
    else:
        nome = random.choice(especialidadeLines)[:-1].title()

    outSql.write("insert into analise values (" +\
        num_analise + ", '" + \
        random.choice(especialidadeLines)[:-1].title() + "', " + \
        tpl[0] + ", " + \
        tpl[1] + ", '" + \
        tpl[2] + "', '" + \
        random.choice(allDatas) + "', '" + \
        nome + "', " +\
        str(random.randint(1, 9)) + ", '" + \
        random.choice(instLines)[:-1] + "');\n"
    )    
instFile.close()




# Venda Farmacia
outSql.write("\n-- Vende_Farmacia\n")
allNumVenda = []
for i in range(1, 450):
    allNumVenda.append(i)
    preco = random.randint(1,50)
    tpl = random.choice(allPrescricao)
    subs = tpl[3]
    quant = tpl[4]
    outSql.write("insert into venda_farmacia values (" + \
        str(i) + ", '" + \
        random.choice(allDatas) + "', '" + \
        subs + "', "+\
        str(quant) + ", " + \
        str(preco) + ", '" + \
        random.choice(allFarmacias) + "');\n")

numVenda = 450
for lst in allArouca:
    numVenda += 1
    subs = "Aspirina"
    quant = random.randint(1,9)
    preco = random.randint(1,50)
    n_venda = str(numVenda)
    data = '2020-11-' + str(random.randint(1, 30))
    # num_cedula, num_doente, data, instituicao
    lst.append(n_venda)
    outSql.write("insert into venda_farmacia values (" + \
    n_venda + ", '" + \
    data + "', '" + \
    subs + "', "+\
    str(quant) + ", " + \
    str(preco) + ", '" + \
    lst[3] + "');\n")



# Prescricao Venda
outSql.write("\n-- Prescricao_venda\n")

for i in range(400):
    tpl = random.choice(allPrescricao)
    allPrescricao.remove(tpl)
    n_venda = random.choice(allNumVenda)
    allNumVenda.remove(n_venda)
    outSql.write("insert into prescricao_venda values (" + \
        tpl[0] + ", " + \
        tpl[1] + ", '" + \
        tpl[2] + "', '" + \
        tpl[3] + "', " + \
        str(n_venda) + ");\n")

for lst in allArouca:
    subs = "Aspirina"
    quant = random.randint(1,9)
    n_venda = random.choice(allNumVenda)
    # num_cedula, num_doente, data, instituicao, numVenda
    outSql.write("insert into prescricao_venda values (" + lst[0] + ", " + lst[1] + ", '" + lst[2] + "', '" + subs + "', " + lst[4] + ");\n")
       



outSql.close()
